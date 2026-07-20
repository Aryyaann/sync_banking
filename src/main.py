from fastapi import FastAPI, HTTPException, Query, Response, Depends
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import create_engine, text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import io
import os
from sync_engine import sync_negocio, DATABASE_URL
from auth import verify_password, create_access_token, get_current_user
from sentence_transformers import SentenceTransformer
from anthropic import Anthropic
import numpy as np
from fastapi import FastAPI, HTTPException, Query, Response, Depends, Header

support_model = SentenceTransformer("paraphrase-multilingual-MiniLM-L12-v2")
anthropic_client = Anthropic()
SUPPORT_TOKEN = os.environ.get("SUPPORT_TOKEN", "WpYQaqZmChnenDJ5N1d9bOURG0gUoTt_")

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://sync-banking-frontend.vercel.app",  # dashboard de clientes
        "http://127.0.0.1:5500",  # tu herramienta de soporte personal (Live Server)
        "http://localhost:5500",
    ],
    allow_methods=["*"],
    allow_headers=["*"],
)

engine = create_engine(DATABASE_URL)


@app.get("/status")
def status():
    return {"status": "vivo"}


@app.post("/auth/login")
def login(email: str = Query(...), password: str = Query(...)):
    with engine.connect() as conn:
        user = conn.execute(text("""
            SELECT id, business_id, password_hash FROM users WHERE email = :email
        """), {"email": email}).mappings().first()

    if not user or not verify_password(password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email o contraseña incorrectos")

    token = create_access_token(str(user["id"]), str(user["business_id"]))
    return {"access_token": token}


@app.get("/transactions")
def get_transactions(current_user=Depends(get_current_user)):
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT fecha, importe, moneda, tipo, contraparte, iban_contraparte,
                   concepto_banco, concepto_detallado, categoria, referencia
            FROM transactions WHERE business_id = :bid ORDER BY fecha DESC
        """), {"bid": current_user["business_id"]}).mappings().all()
    return {"count": len(rows), "transactions": [dict(r) for r in rows]}


@app.post("/sync")
def trigger_sync(current_user=Depends(get_current_user)):
    with engine.connect() as conn:
        conexion = conn.execute(text("""
            SELECT id FROM bank_connections WHERE business_id = :bid AND active = true LIMIT 1
        """), {"bid": current_user["business_id"]}).mappings().first()

    if not conexion:
        raise HTTPException(status_code=404, detail="No hay conexión bancaria activa")

    resultado = sync_negocio(current_user["business_id"], str(conexion["id"]))
    return resultado

@app.get("/support/products")
def support_products(authorization: str = Header(None)):
    if authorization != f"Bearer {SUPPORT_TOKEN}":
        raise HTTPException(status_code=401, detail="No autorizado")
    with engine.connect() as conn:
        rows = conn.execute(text("SELECT DISTINCT product FROM support_chunks ORDER BY product")).fetchall()
    return {"products": [r[0] for r in rows]}


@app.post("/support/ask")
def support_ask(product: str = Query(...), question: str = Query(...), authorization: str = Header(None)):
    if authorization != f"Bearer {SUPPORT_TOKEN}":
        raise HTTPException(status_code=401, detail="No autorizado")

    pregunta_emb = support_model.encode([question])[0]

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT source_file, chunk_text, embedding FROM support_chunks WHERE product = :product
        """), {"product": product}).mappings().all()

    if not rows:
        raise HTTPException(status_code=404, detail=f"No hay documentación indexada para '{product}'")

    similitudes = []
    for row in rows:
        emb = np.array(row["embedding"])
        sim = np.dot(pregunta_emb, emb) / (np.linalg.norm(pregunta_emb) * np.linalg.norm(emb))
        similitudes.append((sim, row["source_file"], row["chunk_text"]))
    similitudes.sort(key=lambda x: x[0], reverse=True)
    relevantes = similitudes[:5]

    contexto = "\n\n---\n\n".join(f"[Fuente: {f}]\n{t}" for _, f, t in relevantes)

    prompt = f"""Eres el centro de soporte interno del producto "{product}". Responde basándote EXCLUSIVAMENTE en el contexto. Si no está la respuesta, dilo.

CONTEXTO:
{contexto}

PREGUNTA: {question}

Responde en español, directo y práctico, citando el archivo fuente."""

    response = anthropic_client.messages.create(
        model="claude-haiku-4-5",
        max_tokens=800,
        messages=[{"role": "user", "content": prompt}],
    )

    return {
        "answer": response.content[0].text,
        "sources": list(set(f for _, f, _ in relevantes)),
    }


@app.get("/transactions/export")
def export_excel(current_user=Depends(get_current_user)):
    import pandas as pd

    with engine.connect() as conn:
        df = pd.read_sql(text("""
            SELECT fecha, tipo, importe, moneda, contraparte, concepto_detallado,
                   concepto_banco, categoria, iban_contraparte, referencia
            FROM transactions WHERE business_id = :bid ORDER BY fecha
        """), conn, params={"bid": current_user["business_id"]})

    if df.empty:
        raise HTTPException(status_code=404, detail="No hay movimientos para este negocio")

    df["categoria"] = df["categoria"].fillna("Sin categoria")
    COLUMNAS = list(df.columns)

    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    ENTRADA_FILL = PatternFill("solid", fgColor="E2EFDA")
    SALIDA_FILL = PatternFill("solid", fgColor="FCE4EC")
    REVISAR_FILL = PatternFill("solid", fgColor="FFF2CC")
    BORDE = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    def limpiar_nombre_hoja(nombre):
        return re.sub(r'[\\/*?:\[\]]', '', str(nombre))[:31] or "Sin categoria"

    def escribir_hoja(ws, data):
        ws.append(COLUMNAS)
        for col in range(1, len(COLUMNAS) + 1):
            c = ws.cell(row=1, column=col)
            c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDE
            c.alignment = Alignment(horizontal="center", vertical="center")
        for _, row in data.iterrows():
            ws.append([row.get(c) for c in COLUMNAS])
            r = ws.max_row
            fill = REVISAR_FILL if row.get("concepto_detallado") == "⚠️ REVISAR" else (
                ENTRADA_FILL if row.get("tipo") == "Entrada" else SALIDA_FILL)
            for col in range(1, len(COLUMNAS) + 1):
                cell = ws.cell(row=r, column=col)
                cell.font = Font(name="Arial", size=10)
                cell.border = BORDE
                cell.fill = fill
                if COLUMNAS[col - 1] == "importe":
                    cell.number_format = '#,##0.00 €;[RED]-#,##0.00 €'
        for col in range(1, len(COLUMNAS) + 1):
            letra = get_column_letter(col)
            max_len = max([len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)], default=10)
            ws.column_dimensions[letra].width = min(max(max_len + 3, 12), 45)
        ws.freeze_panes = "A2"

    wb = openpyxl.Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    escribir_hoja(ws_resumen, df)

    for categoria in sorted(df["categoria"].unique()):
        subset = df[df["categoria"] == categoria]
        nombre = limpiar_nombre_hoja(categoria)
        base, i = nombre, 1
        while nombre in wb.sheetnames:
            i += 1
            nombre = f"{base[:28]}_{i}"
        escribir_hoja(wb.create_sheet(nombre), subset)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte.xlsx"},
    )