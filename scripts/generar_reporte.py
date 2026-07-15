import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

MOVIMIENTOS_PATH = "movimientos_sabadell.xlsx"
REPORTE_PATH = "reporte_sabadell.xlsx"

df = pd.read_excel(MOVIMIENTOS_PATH)
if "categoria" not in df.columns:
    df["categoria"] = "Sin categoria"
df["categoria"] = df["categoria"].fillna("Sin categoria")

COLUMNAS = ["fecha", "tipo", "importe", "moneda", "contraparte",
            "concepto_detallado", "concepto_banco", "categoria",
            "iban_contraparte", "referencia"]
COLUMNAS = [c for c in COLUMNAS if c in df.columns]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
ENTRADA_FILL = PatternFill("solid", fgColor="E2EFDA")   # verde: entra dinero
SALIDA_FILL = PatternFill("solid", fgColor="FCE4EC")    # rojo suave: sale dinero
REVISAR_FILL = PatternFill("solid", fgColor="FFF2CC")   # amarillo: pendiente
BORDE = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

def limpiar_nombre_hoja(nombre):
    nombre = re.sub(r'[\\/*?:\[\]]', '', str(nombre))
    return (nombre or "Sin categoria")[:31]

def escribir_hoja(ws, data):
    ws.append(COLUMNAS)
    for col in range(1, len(COLUMNAS) + 1):
        c = ws.cell(row=1, column=col)
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDE
        c.alignment = Alignment(horizontal="center", vertical="center")

    for _, row in data.iterrows():
        ws.append([row.get(c) for c in COLUMNAS])
        r = ws.max_row
        fill = REVISAR_FILL if row.get("concepto_detallado") == "⚠️ REVISAR" else (
            ENTRADA_FILL if row.get("tipo") == "Entrada" else SALIDA_FILL)
        for col in range(1, len(COLUMNAS) + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(name="Arial", size=10)
            cell.border = BORDE
            cell.fill = fill
            if COLUMNAS[col - 1] == "importe":
                cell.number_format = '#,##0.00 €;[RED]-#,##0.00 €'
            if COLUMNAS[col - 1] == "fecha":
                cell.number_format = 'DD/MM/YYYY'

    if "importe" in COLUMNAS and ws.max_row > 1:
        fila_total = ws.max_row + 1
        col_i = COLUMNAS.index("importe") + 1
        letra = get_column_letter(col_i)
        ws.cell(row=fila_total, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
        t = ws.cell(row=fila_total, column=col_i, value=f"=SUM({letra}2:{letra}{fila_total-1})")
        t.font = Font(name="Arial", bold=True)
        t.number_format = '#,##0.00 €;[RED]-#,##0.00 €'
        t.fill = PatternFill("solid", fgColor="D9E1F2")

    for col in range(1, len(COLUMNAS) + 1):
        letra = get_column_letter(col)
        max_len = max([len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)], default=10)
        ws.column_dimensions[letra].width = min(max(max_len + 3, 12), 45)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{ws.max_row - 1}"

wb = openpyxl.Workbook()
ws_resumen = wb.active
ws_resumen.title = "Resumen"
escribir_hoja(ws_resumen, df.sort_values("fecha"))

for categoria in sorted(df["categoria"].unique()):
    subset = df[df["categoria"] == categoria].sort_values("fecha")
    nombre = limpiar_nombre_hoja(categoria)
    base, i = nombre, 1
    while nombre in wb.sheetnames:
        i += 1
        nombre = f"{base[:28]}_{i}"
    escribir_hoja(wb.create_sheet(nombre), subset)

wb.save(REPORTE_PATH)
print(f"✅ Reporte generado: {REPORTE_PATH}")
print(f"   Hojas: {wb.sheetnames}")